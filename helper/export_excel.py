import openpyxl
from openpyxl.styles import Alignment, Border, Side

def run(formzahl, result, highest, lowest, difference, section1, H, g_360):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Measurement Data'

    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    ws.append([])
    ws.append(['', 'Formzahl', 'Highest', 'Lowest', 'Difference', 'Result'])
    ws.append(['', formzahl, highest, lowest, difference, result])

    ws.append([])

    ws.append(['', 'Konstanta', 'S0', 'M2', 'S2', 'N2', 'K1', 'Q1', 'M4', 'MS4', 'K2', 'P1'])
    h = ['', 'H']
    g = ['', 'g360']

    for i in range(len(H)):
        h.append(round(H[i], 3))
        g.append(round(g_360[i], 3))
    
    ws.append(h)
    ws.append(g)

    for row in ws.iter_rows(min_row=2, max_row=3, min_col=2, max_col=5):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    for row in ws.iter_rows(min_row=5, max_row=7, min_col=2, max_col=12):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    ws.append([])
    ws.append([])

    header = ['', 'JAM']
    header.extend([i for i in range(len(section1[0]))])
    ws.append(header)

    for i in range(len(section1)):
        row = ['', str(i+1)]
        row.extend(section1[i])
        ws.append(row)

    for row in ws.iter_rows(min_row=10, max_row=41, min_col=2, max_col=26):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    return wb
